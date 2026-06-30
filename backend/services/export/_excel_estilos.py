"""Estilos y helpers compartidos del renderer Excel (single + multisheet).

Extraído de _excel para eliminar la duplicación del loop de tabla y las constantes de estilo
que el single y el multisheet repetían, y para mantener cada archivo ≤150 líneas.
"""
from typing import Any, List

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_FILL = PatternFill("solid", fgColor="1e293b")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL = PatternFill("solid", fgColor="f8fafc")
BORDER = Border(
    left=Side(style="thin", color="e2e8f0"),
    right=Side(style="thin", color="e2e8f0"),
    top=Side(style="thin", color="e2e8f0"),
    bottom=Side(style="thin", color="e2e8f0"),
)


def escribir_tabla(ws: Any, val: List[dict], row: int) -> int:
    """Escribe una lista de dicts como tabla estilada desde `row`. Devuelve la fila siguiente."""
    headers = list(val[0].keys())
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h.replace("_", " ").capitalize())
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1
    for i, item in enumerate(val):
        fill = ALT_FILL if i % 2 == 1 else PatternFill()
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=item.get(h, ""))
            cell.font = Font(size=9)
            cell.fill = fill
            cell.border = BORDER
        row += 1
    return row


def autofit(ws: Any) -> None:
    """Ajusta el ancho de cada columna al contenido (máx 60 chars)."""
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
