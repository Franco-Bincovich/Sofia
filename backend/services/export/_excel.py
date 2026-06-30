"""Renderer Excel del motor de export (openpyxl). Genérico: mapea `datos` por estructura.

Hoja única por defecto; si `datos` contiene la clave `_sheets` (dict hoja→datos) genera
multi-hoja. El loop de tabla y los estilos viven en _excel_estilos (deduplicados).
"""
import io
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from services.export._excel_estilos import autofit, escribir_tabla


def build_excel(nombre: str, datos: Dict[str, Any]) -> bytes:
    """Genera un Excel (.xlsx) a partir de (nombre, datos). Despacha a multi-hoja si hay _sheets."""
    sheets = datos.get("_sheets")
    if isinstance(sheets, dict):
        return _build_multisheet(sheets)

    wb = Workbook()
    ws = wb.active
    ws.title = nombre[:31]  # Excel limita a 31 chars el nombre de hoja
    row = 1

    ws.cell(row=row, column=1, value=nombre).font = Font(bold=True, size=13)
    row += 2

    scalars = {k: v for k, v in datos.items()
               if not isinstance(v, (list, dict)) and k not in ("titulo",)}
    if scalars:
        ws.cell(row=row, column=1, value="Resumen").font = Font(bold=True, size=11)
        row += 1
        for key, val in scalars.items():
            ws.cell(row=row, column=1, value=key.replace("_", " ").capitalize()).font = Font(bold=True, size=9)
            ws.cell(row=row, column=2, value=str(val)).font = Font(size=9)
            row += 1
        row += 1

    for key in ("analisis", "contexto_datos"):
        if key in datos and isinstance(datos[key], str):
            ws.cell(row=row, column=1, value=key.replace("_", " ").capitalize()).font = Font(bold=True, size=11)
            row += 1
            cell = ws.cell(row=row, column=1, value=datos[key])
            cell.font = Font(size=9)
            cell.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 80
            row += 2

    for key, val in datos.items():
        if not isinstance(val, list) or not val:
            continue
        ws.cell(row=row, column=1, value=key.replace("_", " ").capitalize()).font = Font(bold=True, size=11)
        row += 1
        if isinstance(val[0], dict):
            row = escribir_tabla(ws, val, row)
        else:
            for item in val:
                ws.cell(row=row, column=1, value=str(item)).font = Font(size=9)
                row += 1
        row += 1

    for key, val in datos.items():
        if not isinstance(val, dict):
            continue
        ws.cell(row=row, column=1, value=key.replace("_", " ").capitalize()).font = Font(bold=True, size=11)
        row += 1
        for k, v in val.items():
            ws.cell(row=row, column=1, value=k).font = Font(bold=True, size=9)
            ws.cell(row=row, column=2, value=str(v)).font = Font(size=9)
            row += 1
        row += 1

    autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_multisheet(sheets: Dict[str, Any]) -> bytes:
    """Genera un workbook con una hoja por clave de `sheets` (escalares + tablas)."""
    wb = Workbook()
    wb.remove(wb.active)  # elimina la hoja vacía por defecto
    for sheet_name, sheet_datos in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        row = 1
        ws.cell(row=row, column=1, value=sheet_name).font = Font(bold=True, size=12)
        row += 2
        scalars = {k: v for k, v in sheet_datos.items() if not isinstance(v, (list, dict))}
        if scalars:
            for k, v in scalars.items():
                ws.cell(row=row, column=1, value=str(k)).font = Font(bold=True, size=9)
                ws.cell(row=row, column=2, value=str(v)).font = Font(size=9)
                row += 1
            row += 1
        for key, val in sheet_datos.items():
            if not isinstance(val, list) or not val:
                continue
            ws.cell(row=row, column=1, value=key.replace("_", " ").capitalize()).font = Font(bold=True, size=11)
            row += 1
            if isinstance(val[0], dict):
                row = escribir_tabla(ws, val, row)
            row += 1
        autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
