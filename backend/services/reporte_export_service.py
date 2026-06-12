"""
Servicio de exportación de reportes a PDF y Excel.
Flujo: router → service → repository → DB
"""
import io
from typing import Any, Dict, Optional
from uuid import UUID

from repositories.reporte_repo import ReporteRepo
from utils.errors import AppError
from utils.logger import logger


class ReporteExportService:
    def __init__(self, repo: Optional[ReporteRepo] = None) -> None:
        self._repo = repo or ReporteRepo()

    def export_pdf(self, reporte_id: UUID) -> bytes:
        """
        Genera un PDF con los datos del reporte indicado.

        Args:
            reporte_id: UUID del reporte a exportar.

        Returns:
            Bytes del PDF generado.

        Raises:
            AppError: REPORTE_NOT_FOUND (404) si el reporte no existe.
            AppError: REPORTE_EXPORT_ERROR (500) si falla la generación.
        """
        reporte = self._repo.find_by_id(str(reporte_id))
        if not reporte:
            raise AppError("Reporte no encontrado", "REPORTE_NOT_FOUND", 404)

        try:
            return _build_pdf(reporte.nombre, reporte.datos)
        except Exception as exc:
            logger.error("Error al generar PDF", extra={"reporte_id": str(reporte_id), "error": str(exc)})
            raise AppError("Error al generar el PDF", "REPORTE_EXPORT_ERROR", 500) from exc

    def export_excel(self, reporte_id: UUID) -> bytes:
        """
        Genera un Excel (.xlsx) con los datos del reporte indicado.

        Args:
            reporte_id: UUID del reporte a exportar.

        Returns:
            Bytes del archivo Excel generado.

        Raises:
            AppError: REPORTE_NOT_FOUND (404) si el reporte no existe.
            AppError: REPORTE_EXPORT_ERROR (500) si falla la generación.
        """
        reporte = self._repo.find_by_id(str(reporte_id))
        if not reporte:
            raise AppError("Reporte no encontrado", "REPORTE_NOT_FOUND", 404)

        try:
            return _build_excel(reporte.nombre, reporte.datos)
        except Exception as exc:
            logger.error("Error al generar Excel", extra={"reporte_id": str(reporte_id), "error": str(exc)})
            raise AppError("Error al generar el Excel", "REPORTE_EXPORT_ERROR", 500) from exc


# ── PDF builder ────────────────────────────────────────────────────────────────

def _build_pdf(nombre: str, datos: Dict[str, Any]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                 fontSize=16, spaceAfter=4, textColor=colors.HexColor("#1e293b"))
    h2_style = ParagraphStyle("h2", parent=styles["Heading2"],
                              fontSize=11, spaceAfter=3, textColor=colors.HexColor("#334155"))
    body_style = ParagraphStyle("body", parent=styles["Normal"],
                                fontSize=9, leading=14, textColor=colors.HexColor("#475569"))
    label_style = ParagraphStyle("label", parent=styles["Normal"],
                                 fontSize=9, leading=14, textColor=colors.HexColor("#64748b"))

    story = [
        Paragraph(nombre, title_style),
        HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e2e8f0"), spaceAfter=8),
    ]

    # ── Datos escalares ────────────────────────────────────────────────────────
    scalars = {k: v for k, v in datos.items()
               if not isinstance(v, (list, dict)) and k not in ("titulo",)}
    if scalars:
        story.append(Paragraph("Resumen", h2_style))
        for key, val in scalars.items():
            label = key.replace("_", " ").capitalize()
            story.append(Paragraph(f"<b>{label}:</b>  {val}", body_style))
        story.append(Spacer(1, 0.4*cm))

    # ── Texto largo (análisis IA) ──────────────────────────────────────────────
    for key in ("analisis", "contexto_datos"):
        if key in datos and isinstance(datos[key], str):
            story.append(Paragraph(key.replace("_", " ").capitalize(), h2_style))
            for line in datos[key].split("\n"):
                story.append(Paragraph(line or "&nbsp;", body_style))
            story.append(Spacer(1, 0.4*cm))

    # ── Tablas ─────────────────────────────────────────────────────────────────
    for key, val in datos.items():
        if not isinstance(val, list) or not val:
            continue
        story.append(Paragraph(key.replace("_", " ").capitalize(), h2_style))
        if isinstance(val[0], dict):
            headers = list(val[0].keys())
            table_data = [[h.replace("_", " ").capitalize() for h in headers]]
            for row in val:
                table_data.append([str(row.get(h, "")) for h in headers])
            col_w = (17 * cm) / len(headers)
            tbl = Table(table_data, colWidths=[col_w] * len(headers))
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(tbl)
        else:
            for item in val:
                story.append(Paragraph(f"• {item}", label_style))
        story.append(Spacer(1, 0.4*cm))

    # ── Dicts simples (excluye claves privadas como _sheets) ──────────────────
    for key, val in datos.items():
        if key.startswith("_") or not isinstance(val, dict):
            continue
        story.append(Paragraph(key.replace("_", " ").capitalize(), h2_style))
        for k, v in val.items():
            story.append(Paragraph(f"<b>{k}:</b>  {v}", body_style))
        story.append(Spacer(1, 0.4*cm))

    doc.build(story)
    return buf.getvalue()


# ── Excel builder ──────────────────────────────────────────────────────────────

def _build_excel(nombre: str, datos: Dict[str, Any]) -> bytes:
    # Despacha a multi-hoja si datos contiene _sheets (ej. anual_consolidado)
    sheets = datos.get("_sheets")
    if isinstance(sheets, dict):
        return _build_excel_multisheet(nombre, sheets)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nombre[:31]  # Excel limita a 31 chars el nombre de hoja

    HEADER_FILL = PatternFill("solid", fgColor="1e293b")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL = PatternFill("solid", fgColor="f8fafc")
    BORDER = Border(
        left=Side(style="thin", color="e2e8f0"),
        right=Side(style="thin", color="e2e8f0"),
        top=Side(style="thin", color="e2e8f0"),
        bottom=Side(style="thin", color="e2e8f0"),
    )

    row = 1

    # Título
    ws.cell(row=row, column=1, value=nombre).font = Font(bold=True, size=13)
    row += 2

    # Escalares
    scalars = {k: v for k, v in datos.items()
               if not isinstance(v, (list, dict)) and k not in ("titulo",)}
    if scalars:
        ws.cell(row=row, column=1, value="Resumen").font = Font(bold=True, size=11)
        row += 1
        for key, val in scalars.items():
            label = key.replace("_", " ").capitalize()
            ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=9)
            ws.cell(row=row, column=2, value=str(val)).font = Font(size=9)
            row += 1
        row += 1

    # Textos largos
    for key in ("analisis", "contexto_datos"):
        if key in datos and isinstance(datos[key], str):
            ws.cell(row=row, column=1, value=key.replace("_", " ").capitalize()).font = Font(bold=True, size=11)
            row += 1
            cell = ws.cell(row=row, column=1, value=datos[key])
            cell.font = Font(size=9)
            cell.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 80
            row += 2

    # Tablas (listas de dicts)
    for key, val in datos.items():
        if not isinstance(val, list) or not val:
            continue
        ws.cell(row=row, column=1, value=key.replace("_", " ").capitalize()).font = Font(bold=True, size=11)
        row += 1

        if isinstance(val[0], dict):
            headers = list(val[0].keys())
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=h.replace("_", " ").capitalize())
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="center")
            row += 1
            for data_row_idx, item in enumerate(val):
                fill = ALT_FILL if data_row_idx % 2 == 1 else PatternFill()
                for col_idx, h in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=item.get(h, ""))
                    cell.font = Font(size=9)
                    cell.fill = fill
                    cell.border = BORDER
                row += 1
        else:
            for item in val:
                ws.cell(row=row, column=1, value=str(item)).font = Font(size=9)
                row += 1
        row += 1

    # Dicts simples
    for key, val in datos.items():
        if not isinstance(val, dict):
            continue
        ws.cell(row=row, column=1, value=key.replace("_", " ").capitalize()).font = Font(bold=True, size=11)
        row += 1
        for k, v in val.items():
            ws.cell(row=row, column=1, value=k).font = Font(bold=True, size=9)
            ws.cell(row=row, column=2, value=str(v)).font = Font(size=9)
            row += 1
        row += 1

    # Ajuste automático de anchos de columna
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Excel multi-hoja ───────────────────────────────────────────────────────────

def _build_excel_multisheet(nombre: str, sheets: Dict[str, Any]) -> bytes:
    """
    Genera un workbook Excel con una hoja por clave de `sheets`.
    Cada valor puede contener escalares (key→value) y listas de dicts (tablas).
    Usado por el informe anual consolidado.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HEADER_FILL = PatternFill("solid", fgColor="1e293b")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL = PatternFill("solid", fgColor="f8fafc")
    BORDER = Border(
        left=Side(style="thin", color="e2e8f0"),
        right=Side(style="thin", color="e2e8f0"),
        top=Side(style="thin", color="e2e8f0"),
        bottom=Side(style="thin", color="e2e8f0"),
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # elimina la hoja vacía por defecto

    for sheet_name, sheet_datos in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        row = 1
        ws.cell(row=row, column=1, value=sheet_name).font = Font(bold=True, size=12)
        row += 2

        scalars = {k: v for k, v in sheet_datos.items() if not isinstance(v, (list, dict))}
        if scalars:
            for k, v in scalars.items():
                ws.cell(row=row, column=1, value=str(k)).font = Font(bold=True, size=9)
                ws.cell(row=row, column=2, value=str(v)).font = Font(size=9)
                row += 1
            row += 1

        for key, val in sheet_datos.items():
            if not isinstance(val, list) or not val:
                continue
            ws.cell(row=row, column=1, value=key.replace("_", " ").capitalize()).font = Font(bold=True, size=11)
            row += 1
            if isinstance(val[0], dict):
                headers = list(val[0].keys())
                for col_idx, h in enumerate(headers, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=h.replace("_", " ").capitalize())
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.border = BORDER
                    cell.alignment = Alignment(horizontal="center")
                row += 1
                for i, item in enumerate(val):
                    fill = ALT_FILL if i % 2 == 1 else PatternFill()
                    for col_idx, h in enumerate(headers, start=1):
                        cell = ws.cell(row=row, column=col_idx, value=item.get(h, ""))
                        cell.font = Font(size=9)
                        cell.fill = fill
                        cell.border = BORDER
                    row += 1
            row += 1

        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
